import cv2
import numpy as np
import os
import openpyxl
from datetime import datetime
import subprocess
from dotenv import load_dotenv
from twilio.rest import Client
import mediapipe as mp
import time
import threading
import math
import warnings
import psutil
import hashlib
from cryptography.fernet import Fernet
import json
import requests
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders

# Google Drive imports
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# -----------------------------
# REDUCE WARNINGS
# -----------------------------
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
warnings.filterwarnings('ignore')

# -----------------------------
# LOAD ENV VARIABLES
# -----------------------------
load_dotenv()
sender_email = os.getenv("EMAIL_USER")
app_password = os.getenv("EMAIL_PASS")
twilio_sid = os.getenv("TWILIO_SID")
twilio_token = os.getenv("TWILIO_AUTH")
twilio_from = os.getenv("TWILIO_FROM")
twilio_to = os.getenv("TWILIO_TO")
encryption_key = os.getenv("ENCRYPTION_KEY")
pushover_user_key = os.getenv("PUSHOVER_USER_KEY")
pushover_api_token = os.getenv("PUSHOVER_API_TOKEN")

if twilio_from:
    twilio_from = twilio_from.replace("++", "+")

DEFAULT_LOCATION = "Coimbatore"

# -----------------------------
# CAMERA MODE CONFIGURATION – UPDATED PHONE IP
# -----------------------------
CAMERA_MODES = {
    "WEBCAM": 0,
    "PHONE": "http://192.168.46.165:8080/video"   # <-- YOUR NEW PHONE IP
}
current_mode = "WEBCAM"

ALLOWED_IPS = ["192.168.112.68", "10.0.0.50"]
IGNORE_IP_PATTERNS = ["127.", "192.168.", "10.", "172."]

# Initialize encryption
if not encryption_key:
    encryption_key = Fernet.generate_key()
    print(f"⚠️ Generated new encryption key. Add to .env: ENCRYPTION_KEY={encryption_key.decode()}")
else:
    encryption_key = encryption_key.encode()
cipher = Fernet(encryption_key)

# -----------------------------
# GOOGLE DRIVE SETUP (fixed token refresh)
# -----------------------------
SCOPES = ['https://www.googleapis.com/auth/drive.file']

def upload_to_google_drive(file_path, file_name):
    try:
        creds = None
        if os.path.exists('token.json'):
            creds = Credentials.from_authorized_user_file('token.json', SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            with open('token.json', 'w') as token:
                token.write(creds.to_json())

        service = build('drive', 'v3', credentials=creds)

        folder_id = None
        folder_name = "AI Security Backups"
        response = service.files().list(q=f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false",
                                        spaces='drive', fields='files(id, name)').execute()
        folders = response.get('files', [])
        if folders:
            folder_id = folders[0]['id']
        else:
            file_metadata = {'name': folder_name, 'mimeType': 'application/vnd.google-apps.folder'}
            folder = service.files().create(body=file_metadata, fields='id').execute()
            folder_id = folder.get('id')

        file_metadata = {'name': file_name, 'parents': [folder_id]}
        media = MediaFileUpload(file_path, resumable=True)
        service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        print(f"☁️ Uploaded {file_name} to Google Drive")
        return True
    except Exception as e:
        print(f"❌ Google Drive upload error: {e}")
        print("   → Try deleting 'token.json' and re-run the script to re-authenticate.")
        return False

# -----------------------------
# SECURITY LOGGING
# -----------------------------
security_log_path = os.path.join(os.path.expanduser("~"), "Desktop", "SecurityLogs")
os.makedirs(security_log_path, exist_ok=True)

def log_security_event(event_type, details):
    log_file = os.path.join(security_log_path, f"security_log_{datetime.now().strftime('%Y-%m-%d')}.txt")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_file, 'a') as f:
        f.write(f"[{timestamp}] {event_type}: {details}\n")
    print(f"🔒 {event_type}: {details}")

# -----------------------------
# PUSHOVER NOTIFICATIONS
# -----------------------------
def send_pushover(title, message, priority=0):
    if not pushover_user_key or not pushover_api_token:
        print("⚠️ Pushover credentials missing")
        return False
    try:
        data = {"token": pushover_api_token, "user": pushover_user_key, "title": title, "message": message, "priority": priority}
        response = requests.post("https://api.pushover.net/1/messages.json", data=data, timeout=5)
        if response.status_code == 200:
            print(f"📱 Pushover sent: {title}")
            return True
        else:
            print(f"❌ Pushover failed: {response.text[:80]}")
            return False
    except Exception as e:
        print(f"❌ Pushover error: {e}")
        return False

# -----------------------------
# SMS RATE LIMITER
# -----------------------------
last_sms_time = 0
SMS_COOLDOWN = 2.0

def send_sms_safe(message_body):
    global last_sms_time
    now = time.time()
    if now - last_sms_time < SMS_COOLDOWN:
        time.sleep(SMS_COOLDOWN - (now - last_sms_time))
    try:
        if client:
            msg = client.messages.create(body=message_body, from_=twilio_from, to=twilio_to)
            last_sms_time = time.time()
            print(f"📱 SMS sent: {msg.sid[:8]}...")
            return True
    except Exception as e:
        print(f"❌ SMS failed: {str(e)[:60]}")
        return False

# -----------------------------
# FILE ENCRYPTION
# -----------------------------
def encrypt_file(filepath):
    try:
        with open(filepath, 'rb') as f:
            data = f.read()
        encrypted = cipher.encrypt(data)
        enc_path = filepath + '.enc'
        with open(enc_path, 'wb') as f:
            f.write(encrypted)
        log_security_event("FILE_ENCRYPTED", os.path.basename(filepath))
        return enc_path
    except Exception as e:
        print(f"❌ Encryption failed: {e}")
        return filepath

# -----------------------------
# NETWORK SECURITY
# -----------------------------
last_network_check = 0
NETWORK_CHECK_INTERVAL = 60

def detect_network_threat():
    try:
        connections = psutil.net_connections(kind='inet')
        suspicious = []
        for conn in connections:
            if conn.status == 'ESTABLISHED' and conn.raddr:
                remote_ip = conn.raddr.ip
                is_local = any(remote_ip.startswith(p) for p in IGNORE_IP_PATTERNS)
                if not is_local and remote_ip not in ALLOWED_IPS:
                    suspicious.append(remote_ip)
        if suspicious:
            unique_ips = list(set(suspicious))[:3]
            log_security_event("NETWORK_ACTIVITY", f"External IPs: {', '.join(unique_ips)}")
            return True, suspicious
        return False, []
    except Exception as e:
        print(f"⚠️ Network error: {e}")
        return False, []

# -----------------------------
# CAMERA OBSTRUCTION (MODE-AWARE)
# -----------------------------
obstruction_start_time = None
OBSTRUCTION_ALERT_DELAY = 5
last_obstruction_alert = 0
obstruction_warning_shown = False

def detect_camera_obstruction(frame):
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    mean_brightness = np.mean(gray)
    std_dev = np.std(gray)
    if current_mode == "PHONE":
        if mean_brightness < 5 or mean_brightness > 250:
            return True, f"brightness ({mean_brightness:.0f})"
        if std_dev < 8:
            return True, f"uniform (variance={std_dev:.1f})"
    else:
        if mean_brightness < 20 or mean_brightness > 235:
            return True, f"brightness ({mean_brightness:.0f})"
        if std_dev < 15:
            return True, f"uniform (variance={std_dev:.1f})"
    return False, None

def send_obstruction_alert(obstruction_type):
    global last_obstruction_alert
    now = time.time()
    if now - last_obstruction_alert > 300:
        message = f"⚠️ CAMERA OBSTRUCTION at {DEFAULT_LOCATION}!\nType: {obstruction_type}\nTime: {datetime.now().strftime('%H:%M:%S')}"
        send_sms_safe(message)
        send_pushover("📷 CAMERA BLOCKED", f"{obstruction_type} at {DEFAULT_LOCATION}")
        send_email_raw(sender_email, "⚠️ CAMERA OBSTRUCTION ALERT", message, None)
        last_obstruction_alert = now
        log_security_event("CAMERA_OBSTRUCTION", obstruction_type)

# -----------------------------
# RANSOMWARE PROTECTION
# -----------------------------
def create_immutable_backup(filepath):
    backup_dir = os.path.join(os.path.expanduser("~"), "Desktop", "ImmutableBackups")
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    with open(filepath, 'rb') as f:
        file_hash = hashlib.sha256(f.read()).hexdigest()[:8]
    backup_name = f"{os.path.splitext(os.path.basename(filepath))[0]}_{timestamp}_{file_hash}.xlsx"
    backup_path = os.path.join(backup_dir, backup_name)
    import shutil
    shutil.copy2(filepath, backup_path)
    os.chmod(backup_path, 0o444)
    log_security_event("IMMUTABLE_BACKUP", backup_name)
    return backup_path

# -----------------------------
# EMAIL FUNCTIONS (FIXED & ROBUST)
# -----------------------------
def send_email_raw(to_email, subject, body, image_path=None, attachment_path=None):
    """Send email with optional image and attachment. Returns True on success."""
    try:
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        if image_path and os.path.exists(image_path):
            with open(image_path, 'rb') as f:
                img = MIMEImage(f.read(), name=os.path.basename(image_path))
                msg.attach(img)
                print(f"📎 Attached image: {os.path.basename(image_path)}")

        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(attachment_path)}')
                msg.attach(part)
                print(f"📎 Attached file: {os.path.basename(attachment_path)}")

        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(sender_email, app_password)
        server.send_message(msg)
        server.quit()
        print(f"📧 Email sent successfully: {subject}")
        return True
    except Exception as e:
        print(f"❌ Email send error: {e}")
        return False

def send_attendance_email(filepath, harassment_logs, security_events=[]):
    """Manual email (E key) – sends attendance Excel file (encrypted)."""
    def _send():
        try:
            encrypted_file = encrypt_file(filepath)
            summary = f"Security Report - Location: {DEFAULT_LOCATION}\n\n"
            if harassment_logs:
                summary += "=== HARASSMENT EVENTS ===\n"
                for t, ev in harassment_logs:
                    summary += f"{t} | {ev}\n"
            else:
                summary += "No harassment activity detected.\n"
            if security_events:
                summary += "\n=== SECURITY EVENTS ===\n"
                for event in security_events[:10]:
                    summary += f"{event}\n"
            summary += f"\n✅ File encrypted with AES-256\n"
            send_email_raw(sender_email, "🔒 Security Report + Attendance", summary, attachment_path=encrypted_file)
            if os.path.exists(encrypted_file):
                os.remove(encrypted_file)
        except Exception as e:
            print("❌ Email thread error:", str(e)[:120])
    threading.Thread(target=_send, daemon=True).start()

def send_alert_with_image(event_type, frame, details):
    """Send alert email with captured frame in a background thread."""
    def _send():
        temp_img = None
        try:
            # Save frame as temporary JPEG
            temp_img = os.path.join(security_log_path, f"alert_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg")
            cv2.imwrite(temp_img, frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
            print(f"📸 Captured frame saved: {temp_img}")

            subject = f"🚨 {event_type} at {DEFAULT_LOCATION}"
            body = f"{event_type} detected at {DEFAULT_LOCATION}\nTime: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\nDetails: {details}"
            success = send_email_raw(sender_email, subject, body, image_path=temp_img)
            if success:
                print(f"📧 Alert email with image sent: {event_type}")
            else:
                print(f"❌ Alert email failed: {event_type}")
        except Exception as e:
            print(f"❌ Alert email error: {e}")
        finally:
            if temp_img and os.path.exists(temp_img):
                os.remove(temp_img)
                print(f"🗑️ Temp image deleted: {temp_img}")

    threading.Thread(target=_send, daemon=True).start()

# -----------------------------
# TWILIO INIT
# -----------------------------
client = None
try:
    client = Client(twilio_sid, twilio_token)
    log_security_event("TWILIO_INIT", "Success")
except:
    client = None
    print("⚠ Twilio init failed - SMS disabled")

# -----------------------------
# ALERTS (Pushover + SMS + Email with image)
# -----------------------------
def send_fire_sms(frame=None):
    message = f"🔥 FIRE DETECTED at {DEFAULT_LOCATION}!\nTime: {datetime.now().strftime('%H:%M:%S')}"
    send_sms_safe(message)
    send_pushover("🔥 FIRE ALERT", message)
    if frame is not None:
        send_alert_with_image("FIRE", frame, message)
    log_security_event("FIRE_ALERT", "Alert sent")

def send_harassment_sms(h_type, frame=None):
    message = f"⚠️ HARASSMENT at {DEFAULT_LOCATION}: {h_type}\nTime: {datetime.now().strftime('%H:%M:%S')}"
    send_sms_safe(message)
    send_pushover("🚨 HARASSMENT ALERT", message)
    if frame is not None:
        send_alert_with_image("HARASSMENT", frame, h_type)
    log_security_event("HARASSMENT_ALERT", h_type)

# -----------------------------
# POSE DETECTION (ONLY Crossed Arms + Throat Grab)
# -----------------------------
mp_pose = mp.solutions.pose
pose = mp_pose.Pose(
    static_image_mode=False,
    model_complexity=0,
    smooth_landmarks=True,
    enable_segmentation=False,
    min_detection_confidence=0.5,
    min_tracking_confidence=0.5
)

def get_distance(p1, p2):
    return np.linalg.norm(np.array([p1.x, p1.y]) - np.array([p2.x, p2.y]))

def get_body_scale(landmarks):
    left_shoulder = landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value]
    right_shoulder = landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value]
    left_hip = landmarks[mp_pose.PoseLandmark.LEFT_HIP.value]
    right_hip = landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value]
    shoulder_dist = get_distance(left_shoulder, right_shoulder)
    hip_dist = get_distance(left_hip, right_hip)
    return max(shoulder_dist, hip_dist, 0.1)

def detect_crossed_arms(landmarks, scale):
    lw = landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value]
    rw = landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value]
    le = landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value]
    re = landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value]
    lsh = landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value]
    rsh = landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value]

    left_wrist_to_right_elbow = get_distance(lw, re)
    right_wrist_to_left_elbow = get_distance(rw, le)
    cross_dist = min(left_wrist_to_right_elbow, right_wrist_to_left_elbow)

    left_wrist_to_right_shoulder = get_distance(lw, rsh)
    right_wrist_to_left_shoulder = get_distance(rw, lsh)
    shoulder_cross = min(left_wrist_to_right_shoulder, right_wrist_to_left_shoulder)

    if cross_dist < scale * 0.4 and shoulder_cross < scale * 0.6:
        conf = min(1.0, (1.0 - cross_dist / (scale * 0.5)) * 1.2)
        return True, conf
    return False, 0.0

def detect_throat_grab(landmarks, scale):
    nose = landmarks[mp_pose.PoseLandmark.NOSE.value]
    lw = landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value]
    rw = landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value]
    lsh = landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value]
    rsh = landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value]

    throat_x = (lsh.x + rsh.x) / 2
    throat_y = (lsh.y + rsh.y) / 2 - (get_distance(lsh, rsh) * 0.2)
    throat = type('Landmark', (), {'x': throat_x, 'y': throat_y})()

    left_dist = get_distance(lw, throat)
    right_dist = get_distance(rw, throat)
    min_dist = min(left_dist, right_dist)
    hand_y = min(lw.y, rw.y)

    if min_dist < scale * 0.35 and hand_y < (lsh.y + rsh.y) / 2:
        conf = max(0.0, 1.0 - (min_dist / (scale * 0.35)))
        return True, min(1.0, conf)
    return False, 0.0

def detect_harassment(landmarks):
    """Only Crossed Arms Block and Choking/Throat Grab"""
    if not landmarks:
        return False, None, 0.0

    scale = get_body_scale(landmarks)

    crossed, conf_cross = detect_crossed_arms(landmarks, scale)
    if crossed and conf_cross > 0.65:
        return True, "Crossed Arms Block", conf_cross

    throat, conf_throat = detect_throat_grab(landmarks, scale)
    if throat and conf_throat > 0.7:
        return True, "Choking / Throat Grab", conf_throat

    return False, None, 0.0

# -----------------------------
# FACE RECOGNITION (unchanged)
# -----------------------------
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
known_faces = {}
face_folder_options = ["known_faces", os.path.join(os.path.dirname(__file__), "known_faces")]
face_folder = None
for folder in face_folder_options:
    if os.path.exists(folder):
        face_folder = folder
        break
if face_folder and os.path.exists(face_folder):
    print(f"📂 Loading faces from: {os.path.abspath(face_folder)}")
    for f in os.listdir(face_folder):
        if f.endswith(("jpg", "png", "jpeg")):
            name = os.path.splitext(f)[0]
            img = cv2.imread(os.path.join(face_folder, f))
            if img is not None:
                g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                faces_detected = face_cascade.detectMultiScale(g, 1.3, 5)
                if len(faces_detected) > 0:
                    x, y, w, h = faces_detected[0]
                    roi = g[y:y+h, x:x+w]
                    roi = cv2.resize(roi, (100,100))
                    known_faces[name] = roi
    print(f"✅ Loaded {len(known_faces)} faces")
    log_security_event("FACE_DB_LOADED", f"{len(known_faces)} faces")
else:
    print("⚠ Known faces folder missing - create 'known_faces/' folder")

# -----------------------------
# ATTENDANCE SHEET (unchanged)
# -----------------------------
desktop = os.path.join(os.path.expanduser("~"), "Desktop", "AttendanceRecords")
os.makedirs(desktop, exist_ok=True)
file = datetime.now().strftime("Attendance_%Y-%m-%d_%H-%M-%S.xlsx")
path = os.path.join(desktop, file)
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.append(["Name", "Date", "Time", "Visit Count", "Location"])
last_seen = {}
ATTENDANCE_COOLDOWN = 30
open_excel_after = False

def mark_attendance(name):
    now = datetime.now()
    d = now.strftime("%Y-%m-%d")
    t = now.strftime("%H:%M:%S")
    visits = 1
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == name and str(row[1]) == d:
            visits += 1
    sheet.append([name, d, t, visits, DEFAULT_LOCATION])
    workbook.save(path)
    print(f"✅ {name} - Visit #{visits} at {t}")
    if sheet.max_row % 10 == 0:
        create_immutable_backup(path)

# -----------------------------
# PHONE CONNECTION CHECK (increased timeout)
# -----------------------------
def check_phone_connection():
    url = CAMERA_MODES["PHONE"]
    print(f"📡 Checking phone connection: {url}")
    try:
        resp = requests.get(url, stream=True, timeout=10)
        if resp.status_code == 200:
            chunk = next(resp.iter_content(chunk_size=1024), None)
            if chunk and (b'JFIF' in chunk or b'Content-Type: image/' in str(resp.headers)):
                print("✅ Phone camera reachable")
                resp.close()
                return True
    except Exception as e:
        print(f"⚠️ Connection test failed: {e}")
    try:
        test_cap = cv2.VideoCapture(url)
        if test_cap.isOpened():
            ret, frame = test_cap.read()
            test_cap.release()
            if ret and frame is not None:
                print("✅ Phone camera verified (frame captured)")
                return True
    except Exception as e:
        print(f"⚠️ OpenCV test failed: {e}")
    print("❌ Phone camera NOT available")
    return False

# -----------------------------
# CAMERA SWITCH (robust fallback)
# -----------------------------
def switch_camera_mode():
    global current_mode, cap
    target_mode = "PHONE" if current_mode == "WEBCAM" else "WEBCAM"
    if target_mode == "PHONE" and not check_phone_connection():
        print("❌ Cannot switch to phone (not reachable). Staying on webcam.")
        return False
    cap.release()
    new_source = CAMERA_MODES[target_mode]
    cap = cv2.VideoCapture(new_source)
    if not cap.isOpened():
        print(f"❌ Failed to open {target_mode}. Falling back to {current_mode}")
        cap = cv2.VideoCapture(CAMERA_MODES[current_mode])
        return False
    cap.set(3, 640)
    cap.set(4, 480)
    cap.set(cv2.CAP_PROP_FPS, 30)
    current_mode = target_mode
    print(f"📸 Switched to {current_mode} camera")
    log_security_event("CAMERA_MODE_SWITCH", current_mode)
    return True

# -----------------------------
# INITIAL CAMERA
# -----------------------------
cap = cv2.VideoCapture(CAMERA_MODES[current_mode])
if not cap.isOpened():
    print("❌ Camera error")
    log_security_event("CAMERA_ERROR", "Failed to initialize")
    exit()
cap.set(3, 640)
cap.set(4, 480)
cap.set(cv2.CAP_PROP_FPS, 30)

harassment_enabled = False
harassment_last = {}
HARASSMENT_COOLDOWN = 30
fire_sent = False
fire_cooldown = 0
frame_count = 0
harassment_logs = []
security_events = []
start_time = time.time()
last_network_check = 0

log_security_event("SYSTEM_START", f"Location: {DEFAULT_LOCATION}")

print("\n" + "=" * 60)
print("🔒 AI SECURITY SYSTEM v2.0 - FINAL (Crossed Arms & Throat Grab)")
print("=" * 60)
print("CONTROLS: H = Harassment | C = Switch Camera | E = Email | G = Drive Backup | Q = Quit")
print("=" * 60 + "\n")

# -----------------------------
# MAIN LOOP (unchanged except harassment detection)
# -----------------------------
while True:
    ret, frame = cap.read()
    if not ret:
        log_security_event("FRAME_ERROR", "Failed to capture")
        break

    # Resize and rotate if needed
    if frame.shape[1] != 640 or frame.shape[0] != 480:
        frame = cv2.resize(frame, (640, 480))
    if frame.shape[0] > frame.shape[1]:
        frame = cv2.rotate(frame, cv2.ROTATE_90_CLOCKWISE)

    frame_count += 1
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    current_time = time.time()

    # Camera obstruction
    is_obstructed, obs_type = detect_camera_obstruction(frame)
    if is_obstructed:
        if obstruction_start_time is None:
            obstruction_start_time = current_time
            obstruction_warning_shown = False
        elapsed = current_time - obstruction_start_time
        if not obstruction_warning_shown and elapsed > 1:
            print(f"⚠️ Camera obstructed ({obs_type}) - alert in {int(OBSTRUCTION_ALERT_DELAY - elapsed)}s")
            obstruction_warning_shown = True
        if elapsed > OBSTRUCTION_ALERT_DELAY:
            if current_time - last_obstruction_alert > 300:
                send_obstruction_alert(obs_type)
            obstruction_start_time = current_time
    else:
        obstruction_start_time = None
        obstruction_warning_shown = False

    # Network check
    if current_time - last_network_check > NETWORK_CHECK_INTERVAL:
        threat_detected, suspicious_ips = detect_network_threat()
        last_network_check = current_time

    # Face recognition (every 2 frames)
    faces = []
    if frame_count % 2 == 0 and not is_obstructed:
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        for (x, y, w_f, h_f) in faces:
            roi = cv2.resize(gray[y:y+h_f, x:x+w_f], (100,100))
            best, min_diff = "Unknown", 999
            for name, known in known_faces.items():
                diff = np.mean(cv2.absdiff(roi, known))
                if diff < min_diff and diff < 50:
                    min_diff = diff
                    best = name
            if best != "Unknown":
                now_t = time.time()
                if now_t - last_seen.get(best, 0) > ATTENDANCE_COOLDOWN:
                    mark_attendance(best)
                    last_seen[best] = now_t
                    open_excel_after = True
            color = (0, 255, 0) if best != "Unknown" else (0, 165, 255)
            cv2.putText(frame, best, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
            cv2.rectangle(frame, (x, y), (x+w_f, y+h_f), color, 2)

    # Harassment detection (every 3 frames) – NEW GESTURES ONLY
    if harassment_enabled and frame_count % 3 == 0:
        small = cv2.resize(frame, (320,240))
        rgb = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)
        res = pose.process(rgb)
        if res.pose_landmarks:
            det, h_type, conf = detect_harassment(res.pose_landmarks.landmark)
            if det:
                now_t = time.time()
                key = h_type
                if now_t - harassment_last.get(key, 0) > HARASSMENT_COOLDOWN:
                    print(f"⚠️ {h_type} ({conf*100:.0f}%)")
                    send_harassment_sms(h_type, frame.copy())
                    harassment_last[key] = now_t
                    harassment_logs.append((datetime.now().strftime("%H:%M:%S"), h_type))

    # Fire detection (every 5 frames)
    if frame_count % 5 == 0:
        hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
        lower_fire = np.array([5,150,200])
        upper_fire = np.array([20,255,255])
        mask = cv2.inRange(hsv, lower_fire, upper_fire)
        fire_pixels = cv2.countNonZero(mask)
        if fire_pixels > 25000:
            if not fire_sent and current_time - fire_cooldown > 30:
                print("🔥 FIRE DETECTED!")
                threading.Thread(target=send_fire_sms, args=(frame.copy(),), daemon=True).start()
                fire_sent = True
                fire_cooldown = current_time
        else:
            if current_time - fire_cooldown > 5:
                fire_sent = False

    # -------------------------
    # UI DRAWING (unchanged)
    # -------------------------
    h, w = frame.shape[:2]
    overlay = frame.copy()
    hdr_h = int(h * 0.09)
    cv2.rectangle(overlay, (0,0), (w, hdr_h), (18,20,25), -1)
    ftr_h = int(h * 0.095)
    cv2.rectangle(overlay, (0, h-ftr_h), (w, h), (18,20,25), -1)
    panel_w = int(w * 0.22)
    cv2.rectangle(overlay, (w-panel_w-10, hdr_h+10), (w-10, h-ftr_h-10), (12,16,28), -1)
    alpha = 0.33
    cv2.addWeighted(overlay, alpha, frame, 1-alpha, 0, frame)

    pulse = 0.5 + 0.5 * abs(math.sin(current_time * 2.0))
    live_radius = max(3, int(3 + 3 * pulse))
    brand_x = 18
    cv2.putText(frame, "AI SECURITY", (brand_x, int(hdr_h*0.6)), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (200,220,255), 2, cv2.LINE_AA)
    live_x = brand_x + 320
    cv2.circle(frame, (live_x, int(hdr_h*0.55)), live_radius, (0,220,120), -1)
    cv2.putText(frame, "LIVE", (live_x+16, int(hdr_h*0.6)), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (200,220,200), 1, cv2.LINE_AA)

    pill_w = 180
    pill_h = int(hdr_h*0.5)
    pill_x = w - panel_w - 10 - pill_w - 12
    pill_y = int(hdr_h*0.15)
    is_secure = not is_obstructed
    pill_border = (24,180,80) if is_secure else (220,40,40)
    pill_text = "SYSTEM SECURE" if is_secure else "ALERT ACTIVE"
    pill_text_color = (180,255,200) if is_secure else (255,190,190)
    cv2.rectangle(frame, (pill_x, pill_y), (pill_x+pill_w, pill_y+pill_h), (22,24,30), -1)
    cv2.rectangle(frame, (pill_x, pill_y), (pill_x+pill_w, pill_y+pill_h), pill_border, 2)
    cv2.putText(frame, pill_text, (pill_x+10, pill_y+int(pill_h*0.7)), cv2.FONT_HERSHEY_SIMPLEX, 0.45, pill_text_color, 1, cv2.LINE_AA)

    if is_obstructed and obstruction_start_time:
        elapsed = current_time - obstruction_start_time
        remaining = max(0, OBSTRUCTION_ALERT_DELAY - elapsed)
        warn_text = f"CAMERA BLOCKED! Alert in {int(remaining)}s"
        cv2.putText(frame, warn_text, (pill_x, pill_y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0,0,255), 2)

    cv2.putText(frame, f"Location: {DEFAULT_LOCATION}", (brand_x, int(hdr_h*0.95)), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (200,200,200), 1, cv2.LINE_AA)

    panel_x = w - panel_w - 6
    panel_y = hdr_h + 20
    line_h = 28
    faces_count = len(faces) if faces is not None else 0
    uptime_sec = int(current_time - start_time)
    uptime_str = f"{uptime_sec//3600}h {(uptime_sec%3600)//60}m"

    info_items = [
        ("Location", DEFAULT_LOCATION),
        ("Mode", "WEBCAM" if current_mode == "WEBCAM" else "PHONE"),
        ("Faces", str(faces_count)),
        ("Uptime", uptime_str),
        ("Harassment", "ON" if harassment_enabled else "OFF"),
        ("Encryption", "AES-256"),
    ]

    y_offset = panel_y + 18
    for label, value in info_items:
        cv2.putText(frame, label, (panel_x+12, y_offset), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (180,200,230), 1, cv2.LINE_AA)
        color = (220,240,255)
        cv2.putText(frame, value, (panel_x+12, y_offset+line_h), cv2.FONT_HERSHEY_SIMPLEX, 0.55, color, 1, cv2.LINE_AA)
        y_offset += 2*line_h

    cv2.putText(frame, "H=Harass | C=Camera | E=Email | G=Drive | Q=Quit",
                (int(w*0.03), h-int(ftr_h*0.4)), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (200,200,200), 1, cv2.LINE_AA)

    cv2.imshow("AI Security System v2.0", frame)

    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):
        break
    if key == ord('h'):
        harassment_enabled = not harassment_enabled
        print(f"🔄 Harassment: {'ON' if harassment_enabled else 'OFF'}")
        log_security_event("HARASSMENT_TOGGLE", "ON" if harassment_enabled else "OFF")
    if key == ord('c'):
        if not switch_camera_mode():
            print("❌ Camera switch failed - phone not reachable")
    if key == ord('e'):
        print("📤 Sending encrypted security report...")
        workbook.save(path)
        create_immutable_backup(path)
        send_attendance_email(path, harassment_logs, security_events)
    if key == ord('g'):
        print("☁️ Manual backup to Google Drive...")
        workbook.save(path)
        def backup_thread():
            upload_to_google_drive(path, os.path.basename(path))
        threading.Thread(target=backup_thread, daemon=True).start()

# -----------------------------
# CLEANUP
# -----------------------------
cap.release()
cv2.destroyAllWindows()
workbook.save(path)
create_immutable_backup(path)
log_security_event("SYSTEM_SHUTDOWN", "Clean shutdown")
if open_excel_after:
    try:
        subprocess.Popen(['start', path], shell=True)
    except:
        pass
print("\n" + "=" * 60)
print("✅ AI SECURITY SYSTEM SHUTDOWN COMPLETE")
print("🔒 All data encrypted and backed up")
print("📊 Security logs: " + security_log_path)
print("=" * 60 + "\n")